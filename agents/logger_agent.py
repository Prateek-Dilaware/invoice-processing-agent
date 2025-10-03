# agents/logger_agent.py

import os
import re
import json
import openpyxl
from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.join("data", "outputs")
EXCEL_FILE = os.path.join(BASE_DIR, "invoices_data.xlsx")
EXTRACTED_FILE = os.path.join(BASE_DIR, "extracted_content.txt")

class LoggerAgent:
    def __init__(self, excel_file=EXCEL_FILE, txt_file=EXTRACTED_FILE):
        self.excel_file = excel_file
        self.txt_file = txt_file
        if not os.path.exists(BASE_DIR):
            os.makedirs(BASE_DIR)
        self._init_excel()

    def _init_excel(self):
        """Create Excel file with required sheets if it doesn’t exist"""
        if not os.path.exists(self.excel_file):
            wb = Workbook()

            # Invoice Summary
            ws1 = wb.active
            ws1.title = "Invoice_Summary"
            ws1.append([
                "DocNo", "DocDt", "SellerGstin", "BuyerGstin", "IRN", "AckNo",
                "EWayBill", "PlaceOfSupply", "Transport", "VehicleNo",
                "ItemCnt(QR)", "TotInvVal(QR)", "TaxableAmount",
                "CGST_Amount", "SGST_Amount", "TotalTax", "VendorName", "ValidationFlag"
            ])

            # Line Items
            ws2 = wb.create_sheet("Line_Items")
            ws2.append([
                "DocNo", "S.No", "Description", "HSN/SAC", "Quantity", "Unit",
                "Rate", "Amount"
            ])

            # QR Meta
            ws3 = wb.create_sheet("QR_Meta")
            ws3.append([
                "DocNo", "SellerGstin", "BuyerGstin", "DocTyp", "DocDt",
                "TotInvVal", "ItemCnt", "MainHsnCode", "IRN", "IrnDt", "RawPayload"
            ])

            wb.save(self.excel_file)

    def _split_invoices(self, text: str):
        """Split text into invoice blocks using 'VALIDATION:' marker"""
        blocks = []
        current = []
        for line in text.splitlines():
            current.append(line)
            if line.strip().startswith("VALIDATION:"):
                blocks.append("\n".join(current))
                current = []
        return blocks

    def _parse_invoice(self, block: str):
        """Parse one invoice block into summary, line items, QR meta"""
        summary, line_items, qr_meta = {}, [], {}

        # --- QR Payload ---
        qr_match = re.search(r"\[--- Decoded QR Payload\(s\) ---\]\s*(\{.*?\})\s*\[", block, re.S)
        if qr_match:
            try:
                data = json.loads(qr_match.group(1)).get("data", {})
                qr_meta = {
                    "DocNo": data.get("DocNo"),
                    "SellerGstin": data.get("SellerGstin"),
                    "BuyerGstin": data.get("BuyerGstin"),
                    "DocTyp": data.get("DocTyp"),
                    "DocDt": data.get("DocDt"),
                    "TotInvVal": data.get("TotInvVal"),
                    "ItemCnt": data.get("ItemCnt"),
                    "MainHsnCode": data.get("MainHsnCode"),
                    "IRN": data.get("Irn"),
                    "IrnDt": data.get("IrnDt"),
                    "RawPayload": json.dumps(data)
                }
                summary.update(qr_meta)
            except Exception as e:
                print(f"⚠️ QR parse error: {e}")

        # --- Extract Header Fields ---
        def safe_search(pattern, text, group=1):
            m = re.search(pattern, text, re.S)
            return m.group(group).strip() if m else None

        def clean_num(val):
            if not val:
                return None
            return float(val.replace(",", ""))

        summary.update({
            "AckNo": safe_search(r"Ack\.No\.\s*:\s*(\d+)", block),
            "EWayBill": safe_search(r"E-Way Bill No\.\s*:\s*(\S+)", block),
            "PlaceOfSupply": safe_search(r"Place of Supply\s*:\s*(.*)", block),
            "VehicleNo": safe_search(r"Vehicle No\.\s*:\s*(\S+)", block),
            "Transport": safe_search(r"Transport\s*:\s*(\S+)", block),
            "ValidationFlag": safe_search(r"(VALIDATION:.*)", block),
        })

        # --- Extract Vendor Name ---
        vendor_name = None
        try:
            lines = block.splitlines()
            for i, line in enumerate(lines):
                if "TAX INVOICE" in line.upper():
                    for j in range(i+1, len(lines)):
                        candidate = lines[j].strip()
                        if candidate:
                            vendor_name = candidate
                            break
                    break
        except Exception as e:
            print(f"⚠️ Vendor name parse error: {e}")

        summary["VendorName"] = vendor_name

        # --- Extract Taxes ---
        tax_match = re.search(
            r"(\d{8})\s+\d+%?\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})(?:\s+([\d,]+\.\d{2}))?",
            block
        )
        if tax_match:
            summary["TaxableAmount"] = clean_num(tax_match.group(2))
            summary["CGST_Amount"] = clean_num(tax_match.group(3))
            summary["SGST_Amount"] = clean_num(tax_match.group(4))
            summary["TotalTax"] = clean_num(tax_match.group(5)) or (
                (summary.get("CGST_Amount") or 0) + (summary.get("SGST_Amount") or 0)
            )

        # --- Robust Multi-line Line Item Parser ---
        items_block = re.search(r"S\.N.*?Amount\(`\s*\)\s*\n(.*?)(?:Add\s*: CGST|HSN/SAC\s+Tax Rate|VALIDATION:)", block, re.S)
        if items_block:
            lines = [l.strip() for l in items_block.group(1).splitlines() if l.strip()]
            for i in range(len(lines)):
                # Look for qty/unit line
                if re.match(r"[\d,\.]+\s+(SET|Pcs\.|KG|Units?)", lines[i]):
                    try:
                        qty_line = lines[i]
                        hsn_line = lines[i - 1]
                        desc_line = lines[i - 2]
                        rate_line = lines[i + 1]
                        amt_line = lines[i + 2]

                        if not re.fullmatch(r"\d{8}", hsn_line):
                            continue

                        s_no = None
                        if re.match(r"^\d+\.", lines[i - 3]):
                            s_no = int(lines[i - 3].split(".")[0])

                        qty, unit = re.match(r"([\d,\.]+)\s+(\S+)", qty_line).groups()
                        line_items.append({
                            "S.No": s_no,
                            "Description": desc_line,
                            "HSN/SAC": int(hsn_line),
                            "Quantity": clean_num(qty),
                            "Unit": unit,
                            "Rate": clean_num(rate_line),
                            "Amount": clean_num(amt_line)
                        })
                    except Exception:
                        continue

        return summary, line_items, qr_meta

    def process(self):
        """Main entry: parse extracted_content.txt and log to Excel"""
        if not os.path.exists(self.txt_file):
            print(f"❌ File not found: {self.txt_file}")
            return

        with open(self.txt_file, "r", encoding="utf-8") as f:
            text = f.read()

        blocks = self._split_invoices(text)
        print(f"Found {len(blocks)} invoices.")

        wb = load_workbook(self.excel_file)
        ws_summary = wb["Invoice_Summary"]
        ws_items = wb["Line_Items"]
        ws_qr = wb["QR_Meta"]

        existing_keys = {f"{row[2]}_{row[0]}" for row in ws_summary.iter_rows(min_row=2, values_only=True)}

        for block in blocks:
            summary, items, qr_meta = self._parse_invoice(block)

            if not summary.get("DocNo") or not summary.get("SellerGstin"):
                continue

            comp_key = f"{summary.get('SellerGstin')}_{summary.get('DocNo')}"
            if comp_key in existing_keys:
                print(f"⚠️ Skipping duplicate {comp_key}")
                continue

            ws_summary.append([
                summary.get("DocNo"), summary.get("DocDt"), summary.get("SellerGstin"),
                summary.get("BuyerGstin"), summary.get("IRN"), summary.get("AckNo"),
                summary.get("EWayBill"), summary.get("PlaceOfSupply"), summary.get("Transport"),
                summary.get("VehicleNo"), summary.get("ItemCnt"), summary.get("TotInvVal"),
                summary.get("TaxableAmount"), summary.get("CGST_Amount"),
                summary.get("SGST_Amount"), summary.get("TotalTax"),
                summary.get("VendorName"),  # <-- added here
                summary.get("ValidationFlag")
            ])

            for item in items:
                ws_items.append([
                    summary.get("DocNo"), item.get("S.No"), item.get("Description"),
                    item.get("HSN/SAC"), item.get("Quantity"), item.get("Unit"),
                    item.get("Rate"), item.get("Amount")
                ])

            if qr_meta:
                ws_qr.append([
                    qr_meta.get("DocNo"), qr_meta.get("SellerGstin"), qr_meta.get("BuyerGstin"),
                    qr_meta.get("DocTyp"), qr_meta.get("DocDt"), qr_meta.get("TotInvVal"),
                    qr_meta.get("ItemCnt"), qr_meta.get("MainHsnCode"), qr_meta.get("IRN"),
                    qr_meta.get("IrnDt"), qr_meta.get("RawPayload")
                ])

            print(f"✅ Logged invoice {summary.get('DocNo')} with {len(items)} items")
            existing_keys.add(comp_key)

        wb.save(self.excel_file)
        wb.close()
        print(f"📁 Data saved to {self.excel_file}")


if __name__ == "__main__":
    agent = LoggerAgent()
    agent.process()
