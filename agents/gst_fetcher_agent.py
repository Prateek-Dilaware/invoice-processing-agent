# agents/gst_fetcher_agent.py

import os
import re
import json
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

BASE_DIR = os.path.join("data", "outputs")
EXCEL_FILE = os.path.join(BASE_DIR, "invoices_data.xlsx")
HSN_MAP_FILE = os.path.join("data", "local_cache", "hsn_gst_map.json")

# --- fallback HSN → GST mapping ---
HSN_FALLBACK = {
    "94035000": {"gst_rate": 18, "description": "Wooden furniture"},
    "94036000": {"gst_rate": 12, "description": "Metal furniture"},
    # Add more as needed...
}


class GSTFetcherAgent:
    def __init__(self, excel_file=EXCEL_FILE, hsn_map_file=HSN_MAP_FILE):
        self.excel_file = excel_file
        self.hsn_map_file = hsn_map_file
        self.hsn_cache = self._load_hsn_cache()

    def _load_hsn_cache(self):
        """Load HSN → GST mapping from JSON file, merged with fallback dict."""
        if os.path.exists(self.hsn_map_file):
            try:
                with open(self.hsn_map_file, "r", encoding="utf-8") as f:
                    return {**HSN_FALLBACK, **json.load(f)}
            except Exception:
                return dict(HSN_FALLBACK)
        return dict(HSN_FALLBACK)

    def _save_hsn_cache(self):
        """Save updated HSN cache to JSON file."""
        os.makedirs(os.path.dirname(self.hsn_map_file), exist_ok=True)
        with open(self.hsn_map_file, "w", encoding="utf-8") as f:
            json.dump(self.hsn_cache, f, indent=4, ensure_ascii=False)

    def _fetch_gst_details(self, hsn: str):
        """
        Fetch GST details (rate + description) for given HSN using VakilSearch.
        No effective date will be fetched/stored.
        """
        # ✅ Step 1: Check cache
        if hsn in self.hsn_cache:
            data = self.hsn_cache[hsn]
            return data.get("gst_rate"), data.get("description"), "Local Cache"

        url = f"https://vakilsearch.com/hsn-code/search/{hsn}"
        gst_rate, description = None, None
        try:
            resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")

                # ✅ Parse table row: HSN | Description | GST%
                for row in soup.find_all("tr"):
                    cols = row.find_all("td")
                    if len(cols) >= 3:
                        hsn_code = cols[0].get_text(strip=True)
                        if hsn_code == hsn:  # exact match
                            description = cols[1].get_text(strip=True) if len(cols) > 1 else None
                            rate_text = cols[2].get_text(strip=True) if len(cols) > 2 else None

                            if rate_text:
                                match = re.search(r"(\d{1,2})", rate_text)
                                if match:
                                    gst_rate = int(match.group(1))
                            break

                # ✅ Save to cache if found
                if gst_rate:
                    self.hsn_cache[hsn] = {
                        "gst_rate": gst_rate,
                        "description": description
                    }
                    return gst_rate, description, url

        except Exception as e:
            print(f"⚠ Error fetching GST for HSN {hsn}: {e}")

        # ✅ Step 2: Fallback dictionary
        if hsn in HSN_FALLBACK:
            data = HSN_FALLBACK[hsn]
            return data["gst_rate"], data.get("description"), "Fallback Dictionary"

        return None, None, "Not Found"

    def process(self):
        if not os.path.exists(self.excel_file):
            print(f"❌ Invoice file not found: {self.excel_file}")
            return

        wb = load_workbook(self.excel_file)
        ws_summary = wb["Invoice_Summary"]
        ws_qr = wb["QR_Meta"]

        # Reset GST_Fetch sheet if exists
        if "GST_Fetch" in wb.sheetnames:
            del wb["GST_Fetch"]
        ws_gst = wb.create_sheet("GST_Fetch")
        ws_gst.append([
            "DocNo", "DocDt", "HSN", "Item Description",
            "GST_Rate (%)", "CGST_Rate (%)", "SGST_Rate (%)", "Source"
        ])

        # Build HSN lookup (DocNo → HSN)
        hsn_lookup = {}
        for row in ws_qr.iter_rows(min_row=2, values_only=True):
            docno, seller, buyer, doctype, docdt, totinv, itemcnt, mainhsn, irn, irndt, raw = row
            if docno and mainhsn:
                hsn_lookup[docno] = str(mainhsn)

        # Process each invoice
        for row in ws_summary.iter_rows(min_row=2, values_only=True):
            docno, docdt, *_ = row
            if not docno:
                continue

            hsn = hsn_lookup.get(docno)
            if not hsn:
                print(f"⚠ No HSN found for {docno}, skipping.")
                continue

            gst_rate, description, source = self._fetch_gst_details(hsn)

            if gst_rate:
                cgst_rate = gst_rate / 2
                sgst_rate = gst_rate / 2
            else:
                cgst_rate = sgst_rate = None

            desc_display = (description[:80] + "...") if description and len(description) > 80 else description

            ws_gst.append([docno, docdt, hsn, desc_display, gst_rate, cgst_rate, sgst_rate, source])
            print(f"✅ GST for {docno}: {gst_rate}% (HSN {hsn} - {desc_display}, Source: {source})")

        wb.save(self.excel_file)
        wb.close()
        self._save_hsn_cache()
        print(f"📁 GST data saved in 'GST_Fetch' sheet of {self.excel_file}'")


if __name__ == "__main__":
    agent = GSTFetcherAgent()
    agent.process()
