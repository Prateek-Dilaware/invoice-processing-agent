# agents/mapper_agent.py

import os
import re
import openpyxl
from openpyxl import load_workbook

BASE_DIR = os.path.join("data", "outputs")
INVOICE_FILE = os.path.join(BASE_DIR, "invoices_data.xlsx")
MASTER_FILE = os.path.join(BASE_DIR, "master_file.xlsx")

class MapperAgent:
    def __init__(self, invoice_file=INVOICE_FILE, master_file=MASTER_FILE):
        self.invoice_file = invoice_file
        self.master_file = master_file

    def _normalize_model(self, text: str) -> str:
        """Normalize model no strings for matching"""
        if not text:
            return ""
        text = text.upper()
        text = text.replace("MODEL NO", "").strip()
        # remove HSN in brackets e.g. (73239920)
        text = re.sub(r"\(\d+\)", "", text)
        # collapse multiple spaces
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def process(self):
        # --- Load Master File ---
        if not os.path.exists(self.master_file):
            print(f"❌ Master file not found: {self.master_file}")
            return

        master_wb = load_workbook(self.master_file)
        master_ws = master_wb.active

        master_map = {}
        for row in master_ws.iter_rows(min_row=2, values_only=True):
            model_no, prod_name, sku, hsn, rate = row
            norm_model = self._normalize_model(str(model_no))
            master_map[norm_model] = {
                "ProductName": prod_name,
                "SKU": sku,
                "HSN": hsn,
                "Rate": rate
            }

        # --- Load Invoice File ---
        if not os.path.exists(self.invoice_file):
            print(f"❌ Invoice file not found: {self.invoice_file}")
            return

        wb = load_workbook(self.invoice_file)
        ws_items = wb["Line_Items"]

        # Reset Mapped_Items and Unmapped_Items sheets if they exist
        for sheet in ["Mapped_Items", "Unmapped_Items"]:
            if sheet in wb.sheetnames:
                del wb[sheet]

        ws_map = wb.create_sheet("Mapped_Items")
        ws_unmap = wb.create_sheet("Unmapped_Items")

        # Headers for mapped
        ws_map.append([
            "DocNo", "S.No", "Description", "HSN (Invoice)", "Quantity", "Unit",
            "Rate (Invoice)", "Amount", "Mapped Model", "Mapped Product Name",
            "Mapped SKU", "Mapped HSN", "Standard Rate", "Rate Match Flag"
        ])
        # Headers for unmapped
        ws_unmap.append([
            "DocNo", "S.No", "Description", "HSN (Invoice)", "Quantity", "Unit",
            "Rate (Invoice)", "Amount", "Normalized Model"
        ])

        # --- Process Line Items ---
        for row in ws_items.iter_rows(min_row=2, values_only=True):
            row = list(row[:8])   # Take only first 8 columns (ignore "Item Category")
            if len(row) < 8:
                row += [None] * (8 - len(row))  # pad missing with None

            docno, sno, desc, hsn, qty, unit, rate, amount = row

            norm_model = self._normalize_model(str(desc))
            mapped = master_map.get(norm_model, None)

            if mapped:
                rate_match = "✅" if float(rate) == float(mapped["Rate"]) else "❌"
                ws_map.append([
                    docno, sno, desc, hsn, qty, unit, rate, amount,
                    norm_model, mapped["ProductName"], mapped["SKU"],
                    mapped["HSN"], mapped["Rate"], rate_match
                ])
            else:
                # Unmapped goes to both sheets
                ws_map.append([
                    docno, sno, desc, hsn, qty, unit, rate, amount,
                    norm_model, "NOT FOUND", "NOT FOUND", "NOT FOUND", "NOT FOUND", "❌"
                ])
                ws_unmap.append([
                    docno, sno, desc, hsn, qty, unit, rate, amount, norm_model
                ])

        wb.save(self.invoice_file)
        wb.close()
        print(f"📁 Mapping completed. Results saved in 'Mapped_Items' and 'Unmapped_Items' sheets of {self.invoice_file}'")


if __name__ == "__main__":
    agent = MapperAgent()
    agent.process()
