# agents/reviewer_agent.py

import os
from openpyxl import load_workbook

BASE_DIR = os.path.join("data", "outputs")
EXCEL_FILE = os.path.join(BASE_DIR, "invoices_data.xlsx")


class ReviewerAgent:
    def __init__(self, excel_file=EXCEL_FILE):
        self.excel_file = excel_file

    def _find_totinv_column(self, ws):
        """Find TotInvVal / Invoice Total column index heuristically."""
        header_cells = list(ws[1])
        header = [cell.value for cell in header_cells]
        header_norm = [str(h).strip().lower() if h else "" for h in header]

        # Strong matches
        for i, h in enumerate(header_norm):
            if not h:
                continue
            if "item" in h or "count" in h or "cnt" in h:
                continue
            if ("totinv" in h) or ("tot inv" in h) or \
               ("total" in h and ("inv" in h or "invoice")) or \
               ("invoice" in h and "total" in h) or \
               ("total" in h and "value" in h) or \
               ("totalinvoice" in h) or ("totinvval" in h):
                return i

        # Looser matches
        for i, h in enumerate(header_norm):
            if not h or "item" in h:
                continue
            if "total" in h and ("value" in h or "amount" in h or "inv" in h or "invoice"):
                return i
            if "qr" in h and ("total" in h or "tot" in h):
                return i

        if len(header_norm) >= 12:
            return 11

        return None

    def _to_float(self, val):
        """Convert a value into float safely."""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(",", "").replace("₹", "").replace("INR", "")
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            try:
                return float(s.split()[0])
            except Exception:
                return None

    def process(self):
        if not os.path.exists(self.excel_file):
            print(f"❌ Invoice file not found: {self.excel_file}")
            return

        wb = load_workbook(self.excel_file)

        ws_items = wb["Line_Items"]
        ws_summary = wb["Invoice_Summary"]
        ws_gst = wb["GST_Fetch"]

        # Reset review sheets if they exist
        for sheet in ["Review_Report", "LineItem_Errors"]:
            if sheet in wb.sheetnames:
                del wb[sheet]

        ws_review = wb.create_sheet("Review_Report")
        ws_errors = wb.create_sheet("LineItem_Errors")

        # Headers
        ws_review.append([
            "DocNo", "Taxable (from Items)", "GST_Rate (%)",
            "CGST_calc", "SGST_calc", "ExpectedTotal",
            "QR_Total (TotInvVal)", "Diff", "InvoiceTotal_OK"
        ])
        ws_errors.append([
            "DocNo", "S.No", "Description", "Quantity", "Rate",
            "Stored Amount", "Computed Amount", "Match Flag"
        ])

        # --- Step 1: Line Item Check ---
        line_totals = {}
        for row in ws_items.iter_rows(min_row=2, values_only=True):
            docno, sno, desc, hsn, qty, unit, rate, amount, *_ = row
            if not docno:
                continue

            try:
                computed = float(qty) * float(rate)
            except Exception:
                computed = None

            try:
                amount_f = float(amount) if amount is not None else None
            except Exception:
                amount_f = None

            match_flag = "✅" if (
                computed is not None and amount_f is not None and
                round(computed, 2) == round(amount_f, 2)
            ) else "❌"

            if match_flag == "❌":
                ws_errors.append([docno, sno, desc, qty, rate, amount, computed, match_flag])

            if docno not in line_totals:
                line_totals[docno] = 0.0
            if amount_f is not None:
                line_totals[docno] += amount_f
            elif computed is not None:
                line_totals[docno] += computed

        # --- Step 2: GST Rate Lookup (from GST_Fetch) ---
        gst_lookup = {}
        for row in ws_gst.iter_rows(min_row=2, values_only=True):
            if len(row) < 8:
                continue
            docno, docdt, hsn, description, gst_rate, cgst_rate, sgst_rate, source = row[:8]
            if docno:
                gst_lookup[docno] = {
                    "HSN": hsn,
                    "Description": description,
                    "Source": source,
                    "GST": float(gst_rate) if gst_rate not in (None, "") else None,
                    "CGST": float(cgst_rate) if cgst_rate not in (None, "") else None,
                    "SGST": float(sgst_rate) if sgst_rate not in (None, "") else None,
                }

        # --- Step 3: Invoice Total Check ---
        summary_lookup = {}
        totinv_idx = self._find_totinv_column(ws_summary)
        if totinv_idx is None:
            print("⚠ Could not auto-detect TotInvVal column.")
        else:
            header_name = ws_summary[1][totinv_idx].value
            print(f"ℹ TotInvVal column index {totinv_idx} (header: '{header_name}')")

        for row in ws_summary.iter_rows(min_row=2, values_only=True):
            docno = row[0]
            if not docno:
                continue
            qr_total = self._to_float(row[totinv_idx]) if totinv_idx is not None else None
            summary_lookup[docno] = qr_total

        # --- Step 4: Compare ---
        for docno, taxable in line_totals.items():
            gst_info = gst_lookup.get(docno)
            qr_total = summary_lookup.get(docno)

            if not gst_info or gst_info.get("CGST") is None or gst_info.get("SGST") is None or qr_total is None:
                print(f"⚠ Skipping {docno}, missing GST info or QR total.")
                continue

            cgst_calc = taxable * (gst_info["CGST"] / 100.0)
            sgst_calc = taxable * (gst_info["SGST"] / 100.0)
            expected_total = taxable + cgst_calc + sgst_calc

            diff = round(expected_total - float(qr_total), 2)
            status = "✅" if abs(diff) < 0.01 else "❌"

            ws_review.append([
                docno, round(taxable, 2), gst_info["GST"],
                round(cgst_calc, 2), round(sgst_calc, 2),
                round(expected_total, 2), float(qr_total), diff, status
            ])

        # --- Step 5: Add "Item Category" column to Line_Items ---
        headers = [cell.value for cell in ws_items[1]]
        if "Item Category" not in headers:
            ws_items.cell(row=1, column=len(headers) + 1, value="Item Category")

        category_col = headers.index("Item Category") + 1 if "Item Category" in headers else len(headers) + 1

        for row_idx, row in enumerate(ws_items.iter_rows(min_row=2, values_only=True), start=2):
            docno = row[0]
            if not docno:
                continue
            gst_info = gst_lookup.get(docno)
            if gst_info:
                ws_items.cell(row=row_idx, column=category_col, value=gst_info.get("Description"))

        wb.save(self.excel_file)
        wb.close()
        print(f"📁 Review done. Results in 'Review_Report', 'LineItem_Errors', and 'Line_Items' updated with 'Item Category'.")


if __name__ == "__main__":
    agent = ReviewerAgent()
    agent.process()